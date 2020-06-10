import openpyxl


def readdata():
    path = "C:\\Users\\Lenovo\\Desktop\\TestData.xlsx"

    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Nagesh"]

    mydict = {}

    for colno in range(1, sheet.max_column + 1):
        colname = sheet.cell(row=1, column=colno).value
        if colname in mydict:
            value = str(mydict[colname])
            newcolname = colname+"_"+value
            mydict[colname] = mydict[colname]+1
            sheet.cell(row=1, column=colno).value = newcolname
        else:
            mydict[colname] = 1

    workbook.save(path)
    workbook.close()


readdata()